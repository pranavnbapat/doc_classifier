from __future__ import annotations

import csv
import json
import os
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import load_workbook

from docint.ingest.models import IngestedAsset


TABULAR_MAX_ROWS = int(os.getenv("TABULAR_MAX_ROWS", "100"))
TABULAR_PREVIEW_ROWS = int(os.getenv("TABULAR_PREVIEW_ROWS", "30"))
XLSX_MAX_SHEETS = int(os.getenv("XLSX_MAX_SHEETS", "10"))
XLSX_MAX_ROWS_PER_SHEET = int(os.getenv("XLSX_MAX_ROWS_PER_SHEET", "25"))
JSON_MAX_RECORDS = int(os.getenv("JSON_MAX_RECORDS", "100"))
JSON_PREVIEW_RECORDS = int(os.getenv("JSON_PREVIEW_RECORDS", "30"))


def _stringify_row(values: List[object]) -> str:
    cleaned = [str(v).strip() for v in values if v is not None and str(v).strip()]
    return " | ".join(cleaned)


def _flatten_json_value(value: Any, prefix: str = "", *, max_depth: int = 2) -> Dict[str, Any]:
    if max_depth < 0:
        return {prefix or "value": value}

    if isinstance(value, dict):
        out: Dict[str, Any] = {}
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            out.update(_flatten_json_value(child, child_prefix, max_depth=max_depth - 1))
        return out

    if isinstance(value, list):
        if not value:
            return {prefix or "value": "[]"}
        if all(not isinstance(item, (dict, list)) for item in value[:8]):
            preview = ", ".join(str(item).strip() for item in value[:8] if str(item).strip())
            return {prefix or "value": preview}
        return {prefix or "value": f"[list:{len(value)}]"}

    return {prefix or "value": value}


def _normalize_json_records(payload: Any) -> List[Dict[str, Any]]:
    if isinstance(payload, list):
        records = payload
    elif isinstance(payload, dict):
        for key in ("records", "items", "data", "results", "rows"):
            value = payload.get(key)
            if isinstance(value, list):
                records = value
                break
        else:
            records = [payload]
    else:
        records = [{"value": payload}]

    normalized: List[Dict[str, Any]] = []
    for item in records[: max(1, JSON_MAX_RECORDS)]:
        if isinstance(item, dict):
            normalized.append(_flatten_json_value(item))
        else:
            normalized.append({"value": item})
    return normalized


def _ordered_columns(records: Iterable[Dict[str, Any]]) -> List[str]:
    columns: List[str] = []
    seen = set()
    for record in records:
        for key in record.keys():
            if key not in seen:
                seen.add(key)
                columns.append(key)
    return columns


def ingest_delimited_file(file_path: str, filename: str, delimiter: str) -> IngestedAsset:
    path = Path(file_path)
    rows: List[List[str]] = []
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for idx, row in enumerate(reader):
            rows.append(row)
            if idx >= max(1, TABULAR_MAX_ROWS) - 1:
                break

    lines: List[str] = []
    if rows:
        header = _stringify_row(rows[0])
        if header:
            lines.append(f"Columns: {header}")
        for idx, row in enumerate(rows[1: max(1, TABULAR_PREVIEW_ROWS) + 1], start=1):
            rendered = _stringify_row(row)
            if rendered:
                lines.append(f"Row {idx}: {rendered}")

    text = "\n".join(lines).strip()
    suffix = path.suffix.lower()
    mime = "text/csv" if suffix == ".csv" else "text/tab-separated-values"

    return IngestedAsset(
        asset_type=suffix.lstrip("."),
        filename=filename,
        source_path=file_path,
        text=text,
        lines=lines,
        units=max(1, len(rows) - 1 if rows else 1),
        unit_label="rows",
        source="tabular_text",
        mime_type=mime,
        visual_candidate=False,
        ocr_supported=False,
        meta={
            "delimiter": delimiter,
            "preview_rows": max(0, len(rows) - 1),
            "column_count": len(rows[0]) if rows else 0,
        },
    )


def ingest_xlsx(file_path: str, filename: str) -> IngestedAsset:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    lines: List[str] = []
    total_rows = 0
    max_columns = 0
    sheet_names = workbook.sheetnames

    for sheet_name in sheet_names[: max(1, XLSX_MAX_SHEETS)]:
        ws = workbook[sheet_name]
        lines.append(f"Sheet: {sheet_name}")
        preview_count = 0
        for row in ws.iter_rows(values_only=True):
            max_columns = max(max_columns, len([cell for cell in row if cell is not None and str(cell).strip()]))
            rendered = _stringify_row(list(row))
            if rendered:
                label = "Columns" if preview_count == 0 else f"Row {preview_count}"
                lines.append(f"{label}: {rendered}")
                preview_count += 1
                total_rows += 1
            if preview_count >= max(1, XLSX_MAX_ROWS_PER_SHEET):
                break

    text = "\n".join(lines).strip()

    return IngestedAsset(
        asset_type="xlsx",
        filename=filename,
        source_path=file_path,
        text=text,
        lines=lines,
        units=max(1, total_rows),
        unit_label="rows",
        source="xlsx_text",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        visual_candidate=False,
        ocr_supported=False,
        meta={
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
            "preview_rows": total_rows,
            "max_columns": max_columns,
        },
    )


def ingest_json(file_path: str, filename: str) -> IngestedAsset:
    path = Path(file_path)
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        payload = json.load(handle)

    records = _normalize_json_records(payload)
    columns = _ordered_columns(records)

    lines: List[str] = []
    if columns:
        lines.append(f"Columns: {_stringify_row(columns)}")
    for idx, record in enumerate(records[: max(1, JSON_PREVIEW_RECORDS)], start=1):
        row = [record.get(column, "") for column in columns]
        rendered = _stringify_row(row)
        if rendered:
            lines.append(f"Row {idx}: {rendered}")

    text = "\n".join(lines).strip()

    return IngestedAsset(
        asset_type="json",
        filename=filename,
        source_path=file_path,
        text=text,
        lines=lines,
        units=max(1, len(records)),
        unit_label="rows",
        source="json_text",
        mime_type="application/json",
        visual_candidate=False,
        ocr_supported=False,
        meta={
            "preview_rows": len(records),
            "column_count": len(columns),
            "top_level_type": type(payload).__name__,
            "record_count": len(records),
        },
    )
